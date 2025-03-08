import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Tamaño del mundo
FILAS = 24
COLUMNAS = 24

# Tipos de resource_tiles
RESOURCE_TILES = ["water", "rock", "soil", "soil+"]

# Colores para los resource_tiles
COLORES = {
    "water": "00B0F0",  # Azul claro
    "rock": "808080",   # Gris
    "soil": "964B00",   # Marrón
    "soil+": "00FF00"   # Verde
}

# Colores para el texto según el tipo de entidad
COLORES_TEXTO = {
    "Plant": "000000",  # Negro
    "Animal": "FFFFFF",  # Blanco
    "Fungi": "FF0000",   # Rojo
}

# Inicializar el mundo
def inicializar_mundo():
    mundo = [[random.choice(RESOURCE_TILES) for _ in range(COLUMNAS)] for _ in range(FILAS)]
    
    # Asegurar al menos una celda de cada tipo
    mundo[0][0] = "water"
    mundo[0][1] = "rock"
    mundo[0][2] = "soil+"
    
    return mundo

class Entidad:
    def __init__(self, tipo, subtipo, fila, columna, salud, energia, ticks_vida, ticks_sin_comer=0, nacido_en_agua=False):
        self.tipo = tipo
        self.subtipo = subtipo
        self.estado = "live"
        self.fila = fila
        self.columna = columna
        self.salud = salud
        self.energia = energia
        self.ticks_vida = ticks_vida
        self.ticks_sin_comer = ticks_sin_comer
        self.nacido_en_agua = nacido_en_agua

    def morir(self):
        self.estado = "dead"

    def eliminar(self):
        self.estado = "remove"

class Animal(Entidad):
    def __init__(self, subtipo, fila, columna, nacido_en_agua=False):
        vida = 24 + random.randint(-1, 1)  # Vida base de 24 ticks con variación de ±1
        super().__init__("Animal", subtipo, fila, columna, 100, 100, vida, 0, nacido_en_agua)
        self.ticks_existencia = 0

    def mover(self, mundo):
        if self.estado == "dead":
            return

        # Movimiento aleatorio entre 1 y 3 celdas
        movimiento = random.choice([(0, 1), (0, -1), (1, 0), (-1, 0), (1, 1), (1, -1), (-1, 1), (-1, -1)])
        distancia = random.randint(1, 3)
        nueva_fila = self.fila + movimiento[0] * distancia
        nueva_columna = self.columna + movimiento[1] * distancia

        # Asegurar que el movimiento esté dentro de los límites
        nueva_fila = max(0, min(FILAS - 1, nueva_fila))
        nueva_columna = max(0, min(COLUMNAS - 1, nueva_columna))

        # Verificar si el animal puede moverse a la nueva posición
        if mundo[nueva_fila][nueva_columna] == "water" and not self.nacido_en_agua:
            return  # No puede moverse si está en agua y no nació en agua

        # Permitir que los animales estén en la misma celda que plant-low, pero no en plant-high
        ocupada = any(entidad.fila == nueva_fila and entidad.columna == nueva_columna for entidad in entidades if entidad.estado == "live" and (entidad.subtipo != "plant-low"))
        if not ocupada:
            self.fila = nueva_fila
            self.columna = nueva_columna
            self.ticks_existencia += 1

    def consumir(self, mundo, entidades):
        if self.estado == "dead":
            return

        # Lógica de consumo según el subtipo
        celdas_adyacentes = [(self.fila + i, self.columna + j) for i in range(-1, 2) for j in range(-1, 2)
                             if 0 <= self.fila + i < FILAS and 0 <= self.columna + j < COLUMNAS]
        objetivos = [entidad for entidad in entidades if (entidad.fila, entidad.columna) in celdas_adyacentes
                     and entidad.estado == "live" and (entidad.tipo == "Plant" or entidad.tipo == "Fungi")]

        if objetivos:
            objetivo = random.choice(objetivos)
            if self.subtipo == "animal-big":
                if objetivo.tipo == "Plant":
                    if objetivo.subtipo == "plant-high":
                        objetivo.salud -= 1
                    else:
                        objetivo.salud -= 2
            elif self.subtipo == "animal-small":
                if objetivo.tipo == "Plant" and objetivo.subtipo == "plant-low":
                    objetivo.salud -= 1
                elif objetivo.tipo == "Fungi":
                    objetivo.salud -= 1
            
            # Verificar si la salud del objetivo llega a 0
            if objetivo.salud <= 0:
                objetivo.morir()

            self.ticks_sin_comer = 0
        else:
            self.ticks_sin_comer += 1

        # Verificar si el animal ha alcanzado su vida máxima
        if self.ticks_existencia >= self.ticks_vida:
            self.morir()  # Muerte por alcanzar la vida máxima

        if self.ticks_sin_comer > 3:
            self.morir()

        if self.ticks_existencia >= 12 and self.subtipo == "animal-small":
            self.subtipo = "animal-big"

        # Generación de nuevo animal
        if self.ticks_existencia % 6 == 0:
            vecinos = [entidad for entidad in entidades if entidad.estado == "live" and entidad.fila in [self.fila - 1, self.fila, self.fila + 1]
                       and entidad.columna in [self.columna - 1, self.columna, self.columna + 1]]
            if any(isinstance(vecino, Animal) and vecino.subtipo in ["animal-big", "animal-small"] for vecino in vecinos):
                if random.random() < 0.5:
                    entidades.append(Animal("animal-small", *generar_posicion_aleatoria(mundo, entidades)))

    def morir(self):
        self.estado = "dead"
        self.salud = 0  # Reducir la salud a 0 al morir

class Fungi(Entidad):
    def __init__(self, fila, columna):
        super().__init__("Fungi", None, fila, columna, 3, 0, 0)

    def actuar(self, mundo, entidades):
        entidad_muerta = next((entidad for entidad in entidades if entidad.fila == self.fila and entidad.columna == self.columna and entidad.estado == "dead" and entidad.tipo != "Animal"), None)
        if entidad_muerta:
            if self.ticks_vida % 2 == 0:
                self.salud += 1
            if random.random() < 0.2:
                entidad_muerta.eliminar()
                mundo[self.fila][self.columna] = "soil"
        self.ticks_vida += 1

        if self.salud <= 0:
            self.eliminar()


class Planta(Entidad):
    def __init__(self, subtipo, fila, columna):
        super().__init__("Plant", subtipo, fila, columna, 5, 0, 0)

    def crecer(self, mundo):
        if self.subtipo == "plant-low":
            if self.ticks_vida >= 24 or (self.ticks_vida >= 16 and mundo[self.fila][self.columna] == "soil+"):
                self.subtipo = "plant-high"
        
        self.ticks_vida += 1

        # Verificar si la salud llega a 0
        if self.salud <= 0:
            self.morir()

        # Generar una nueva planta cada 8 ticks con un 30% de probabilidad
        if self.ticks_vida % 8 == 0:
            if random.random() < 0.3:  # 30% de probabilidad
                # Intentar colocar una nueva plant-low en una celda adyacente
                for i in range(-1, 2):
                    for j in range(-1, 2):
                        if (i != 0 or j != 0):  # No considerar la celda actual
                            nueva_fila = self.fila + i
                            nueva_columna = self.columna + j
                            # Verificar límites
                            if 0 <= nueva_fila < FILAS and 0 <= nueva_columna < COLUMNAS:
                                # Verificar que la celda está vacía y no tiene plantas
                                if mundo[nueva_fila][nueva_columna] != "rock" and not any(
                                    entidad.fila == nueva_fila and entidad.columna == nueva_columna and
                                    (entidad.subtipo == "plant-low" or entidad.subtipo == "plant-high")
                                    for entidad in entidades
                                ):
                                    # Agregar la nueva planta
                                    entidades.append(Planta("plant-low", nueva_fila, nueva_columna))
                                    break  # Salir del bucle después de generar una planta

        if self.ticks_vida >= 30 + random.randint(-3, 3):
            self.morir()


# Función para generar entidades aleatorias
def generar_entidades_aleatorias(mundo):
    entidades = []

    # Asegurar al menos una planta pequeña y un animal pequeño
    entidades.append(Planta("plant-low", *generar_posicion_aleatoria(mundo, entidades)))
    entidades.append(Animal("animal-small", *generar_posicion_aleatoria(mundo, entidades)))

    # Generar un número aleatorio de plantas, animales y hongos
    num_plantas = int(input("Ingrese número de plantas: "))
    num_animales = int(input("Ingrese número de animales: "))
    num_hongos = int(input("Ingrese número de hongos: "))

    for _ in range(num_plantas):
        while True:
            fila, columna = generar_posicion_aleatoria(mundo, entidades)
            # Verificar que no esté en una celda con rock
            if mundo[fila][columna] != "rock":
                entidades.append(Planta(random.choice(["plant-low", "plant-high"]), fila, columna))
                break

    for _ in range(num_animales):
        subtipo = random.choice(["animal-small", "animal-big"])
        entidades.append(Animal(subtipo, *generar_posicion_aleatoria(mundo, entidades)))

    for _ in range(num_hongos):
        entidades.append(Fungi(*generar_posicion_aleatoria(mundo, entidades)))

    return entidades


# Función para generar una posición aleatoria no ocupada
def generar_posicion_aleatoria(mundo, entidades):
    while True:
        fila = random.randint(0, FILAS - 1)
        columna = random.randint(0, COLUMNAS - 1)
        # Verificar que la celda no esté ocupada por otra entidad
        if not any(entidad.fila == fila and entidad.columna == columna for entidad in entidades):
            return fila, columna


# Función para simular un tick
def simular_tick(mundo, entidades, tick, hoja):
    if tick == 0:
        # En el tick 0, solo se representa la ubicación de las entidades
        # No se hace nada más
        guardar_estado_en_excel(mundo, entidades, tick, hoja)
        return

    for entidad in entidades:
        if entidad.estado == "live":
            if isinstance(entidad, Animal):
                entidad.mover(mundo)
                entidad.consumir(mundo, entidades)
            elif isinstance(entidad, Planta):
                entidad.crecer(mundo)
            elif isinstance(entidad, Fungi):
                entidad.actuar(mundo, entidades)

    # Generar nuevos Fungi en celdas con entidades muertas
    for entidad in entidades:
        if entidad.estado == "dead" and random.random() < 0.1:
            entidades.append(Fungi(entidad.fila, entidad.columna))

    # Guardar el estado del mundo en el archivo XLS
    guardar_estado_en_excel(mundo, entidades, tick, hoja)


# Función para guardar el estado en el archivo XLS
def guardar_estado_en_excel(mundo, entidades, tick, hoja):
    # Configurar el estilo de las celdas
    hoja.title = f"Tick {tick + 1}"
    for fila in range(FILAS):
        for columna in range(COLUMNAS):
            celda = hoja.cell(row=fila + 1, column=columna + 1)
            resource_tile = mundo[fila][columna]
            celda.fill = PatternFill(start_color=COLORES[resource_tile], end_color=COLORES[resource_tile], fill_type="solid")
            
            # Agregar entidades en la celda (solo las que no están en estado "remove")
            entidades_en_celda = [
                entidad
                for entidad in entidades
                if entidad.fila == fila and entidad.columna == columna and entidad.estado != "remove"
            ]
            if entidades_en_celda:
                texto = []
                for entidad in entidades_en_celda:
                    texto.append(
                        f"{entidad.tipo} ({entidad.subtipo if entidad.subtipo else ''})\n"
                        f"Salud: {entidad.salud}\n"
                        f"Estado: {entidad.estado}"
                    )
                # Aplicar texto en blanco
                celda.font = Font(color="FFFFFF")  # Texto en blanco
                celda.value = "\n".join(texto)
                celda.alignment = Alignment(vertical="top")

                # Aplicar el color de texto según el tipo de entidad
                celda.font = Font(color=COLORES_TEXTO[entidades_en_celda[0].tipo])  # Usar el color del primer tipo
                celda.value = "\n".join(texto)
                celda.alignment = Alignment(vertical="top")


# Inicializar el mundo y las entidades
mundo = inicializar_mundo()
entidades = generar_entidades_aleatorias(mundo)

# Crear un archivo Excel
libro = Workbook()
hoja = libro.active

# Solicitar al usuario el número de ticks
try:
    num_ticks = int(input("Ingrese el número de ticks a simular: "))
except ValueError:
    print("Entrada inválida. Se usará un valor predeterminado de 10 ticks.")
    num_ticks = 10

# Ejecutar la simulación
for tick in range(num_ticks):
    if tick > 0:
        hoja = libro.create_sheet(title=f"Tick {tick + 1}")
    simular_tick(mundo, entidades, tick, hoja)

# Guardar el archivo Excel
libro.save("simulacion.xlsx")
print("Simulación completada. Los resultados se han guardado en 'simulacion.xlsx'.")