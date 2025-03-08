#vida.py
import pygame
import random
from animal import AnimalSmall, AnimalBig
from plant import PlantLow, PlantHigh
from fungi import Fungi
from openpyxl import Workbook  # Para generar el archivo Excel
 
# Configuración de Pygame
pygame.init()
SCREEN_SIZE = (1024, 1024)  # Tamaño de la ventana
CELL_SIZE = SCREEN_SIZE[0] // 24  # Tamaño de cada celda (24x24 grid)
screen = pygame.display.set_mode(SCREEN_SIZE)
clock = pygame.time.Clock()

# Cargar imágenes
fungi_img = pygame.image.load("fungi.png")
animal_small_img = pygame.image.load("animal1.png")
animal_big_img = pygame.image.load("animal2.png")
plant_low_img = pygame.image.load("plant1.png")
plant_high_img = pygame.image.load("plant2.png")
skull_img = pygame.image.load("skull.png")  # Imagen para entidades muertas

# Escalar imágenes al tamaño de las celdas
fungi_img = pygame.transform.scale(fungi_img, (CELL_SIZE, CELL_SIZE))
animal_small_img = pygame.transform.scale(animal_small_img, (CELL_SIZE, CELL_SIZE))
animal_big_img = pygame.transform.scale(animal_big_img, (CELL_SIZE, CELL_SIZE))
plant_low_img = pygame.transform.scale(plant_low_img, (CELL_SIZE, CELL_SIZE))
plant_high_img = pygame.transform.scale(plant_high_img, (CELL_SIZE, CELL_SIZE))
skull_img = pygame.transform.scale(skull_img, (CELL_SIZE, CELL_SIZE))

# Colores para recursos (rock, soil, soil+, water)
COLORS = {
    "ground": (34, 139, 34),  # Verde para el suelo
    "water": (0, 0, 255),  # Azul para agua
    "rock": (128, 128, 128),  # Gris para rocas
    "soil": (139, 69, 19),  # Marrón para tierra
    "soil+": (210, 105, 30),  # Marrón oscuro para tierra fértil
}

class World:
    def __init__(self, width=24, height=24):
        self.width = width
        self.height = height
        self.grid = [[None for _ in range(width)] for _ in range(height)]
        self.entities = []
        self.resource_tiles = []  # Almacena casillas de recursos (rock, soil, soil+, water)

        # Generar características mínimas iniciales
        self.generate_initial_resources()

    def generate_initial_resources(self):
        """Genera al menos una casilla de cada tipo: rock, soil, soil+, water."""
        # Generar una casilla de roca
        x, y = random.randint(0, self.width - 1), random.randint(0, self.height - 1)
        self.grid[y][x] = "rock"
        self.resource_tiles.append(("rock", x, y))

        # Generar una casilla de tierra
        while True:
            x, y = random.randint(0, self.width - 1), random.randint(0, self.height - 1)
            if self.grid[y][x] is None:
                self.grid[y][x] = "soil"
                self.resource_tiles.append(("soil", x, y))
                break

        # Generar una casilla de tierra fértil (soil+)
        while True:
            x, y = random.randint(0, self.width - 1), random.randint(0, self.height - 1)
            if self.grid[y][x] is None:
                self.grid[y][x] = "soil+"
                self.resource_tiles.append(("soil+", x, y))
                break

        # Generar una casilla de agua
        while True:
            x, y = random.randint(0, self.width - 1), random.randint(0, self.height - 1)
            if self.grid[y][x] is None:
                self.grid[y][x] = "water"
                self.resource_tiles.append(("water", x, y))
                break

    def add_entity(self, entity):
        """Añade una entidad a la simulación si la celda está vacía."""
        if 0 <= entity.x < self.width and 0 <= entity.y < self.height and self.grid[entity.y][entity.x] is None:
            self.grid[entity.y][entity.x] = entity
            self.entities.append(entity)

    def update(self):
        """Ejecuta un tick de la simulación."""
        for entity in list(self.entities):
            if entity.state == 'remove':
                self.remove_entity(entity)
                continue

            if entity.state == 'dead':
                continue  # Entidades muertas no hacen nada

            if isinstance(entity, (AnimalSmall, AnimalBig)):
                entity.move(self.grid)
                entity.consume_plant(self.grid)
                entity.check_death()
                entity.check_reproduction(self.grid)
            elif isinstance(entity, (PlantLow, PlantHigh)):
                entity.grow(self.grid)
                entity.check_death()
            elif isinstance(entity, Fungi):
                entity.grow(self.grid)
                entity.check_death()

        # Generar hongos en cadáveres
        for x in range(self.width):
            for y in range(self.height):
                cell = self.grid[y][x]
                if hasattr(cell, 'state') and cell.state == 'dead' and random.random() < 0.1:
                    self.add_entity(Fungi(x, y))

    def remove_entity(self, entity):
        """Elimina una entidad de la simulación."""
        if entity in self.entities:
            self.entities.remove(entity)
        if self.grid[entity.y][entity.x] == entity:
            self.grid[entity.y][entity.x] = None

    def display(self):
        """Dibuja el mundo en la ventana de Pygame."""
        screen.fill(COLORS["ground"])  # Fondo verde (suelo)

        for y in range(self.height):
            for x in range(self.width):
                cell = self.grid[y][x]
                if cell == "rock":
                    color = COLORS["rock"]
                    pygame.draw.rect(screen, color, (x * CELL_SIZE, y * CELL_SIZE, CELL_SIZE, CELL_SIZE))
                elif cell == "soil":
                    color = COLORS["soil"]
                    pygame.draw.rect(screen, color, (x * CELL_SIZE, y * CELL_SIZE, CELL_SIZE, CELL_SIZE))
                elif cell == "soil+":
                    color = COLORS["soil+"]
                    pygame.draw.rect(screen, color, (x * CELL_SIZE, y * CELL_SIZE, CELL_SIZE, CELL_SIZE))
                elif cell == "water":
                    color = COLORS["water"]
                    pygame.draw.rect(screen, color, (x * CELL_SIZE, y * CELL_SIZE, CELL_SIZE, CELL_SIZE))
                elif isinstance(cell, PlantLow) and cell.state != 'remove':
                    if cell.state == 'dead':
                        screen.blit(skull_img, (x * CELL_SIZE, y * CELL_SIZE))
                    else:
                        screen.blit(plant_low_img, (x * CELL_SIZE, y * CELL_SIZE))
                elif isinstance(cell, PlantHigh) and cell.state != 'remove':
                    if cell.state == 'dead':
                        screen.blit(skull_img, (x * CELL_SIZE, y * CELL_SIZE))
                    else:
                        screen.blit(plant_high_img, (x * CELL_SIZE, y * CELL_SIZE))
                elif isinstance(cell, AnimalSmall) and cell.state != 'remove':
                    if cell.state == 'dead':
                        screen.blit(skull_img, (x * CELL_SIZE, y * CELL_SIZE))
                    else:
                        screen.blit(animal_small_img, (x * CELL_SIZE, y * CELL_SIZE))
                elif isinstance(cell, AnimalBig) and cell.state != 'remove':
                    if cell.state == 'dead':
                        screen.blit(skull_img, (x * CELL_SIZE, y * CELL_SIZE))
                    else:
                        screen.blit(animal_big_img, (x * CELL_SIZE, y * CELL_SIZE))
                elif isinstance(cell, Fungi) and cell.state != 'remove':
                    if cell.state == 'dead':
                        screen.blit(skull_img, (x * CELL_SIZE, y * CELL_SIZE))
                    else:
                        screen.blit(fungi_img, (x * CELL_SIZE, y * CELL_SIZE))

        pygame.display.update()  # Actualizar la pantalla

    def get_grid_state(self):
        """Devuelve el estado actual de la cuadrícula como una lista de listas."""
        grid_state = []
        for y in range(self.height):
            row = []
            for x in range(self.width):
                cell = self.grid[y][x]
                if cell == "rock":
                    row.append("rock")
                elif cell == "soil":
                    row.append("soil")
                elif cell == "soil+":
                    row.append("soil+")
                elif cell == "water":
                    row.append("water")
                elif isinstance(cell, PlantLow):
                    row.append("PlantLow" if cell.state == 'live' else "PlantLow (dead)")
                elif isinstance(cell, PlantHigh):
                    row.append("PlantHigh" if cell.state == 'live' else "PlantHigh (dead)")
                elif isinstance(cell, AnimalSmall):
                    row.append("AnimalSmall" if cell.state == 'live' else "AnimalSmall (dead)")
                elif isinstance(cell, AnimalBig):
                    row.append("AnimalBig" if cell.state == 'live' else "AnimalBig (dead)")
                elif isinstance(cell, Fungi):
                    row.append("Fungi" if cell.state == 'live' else "Fungi (dead)")
                else:
                    row.append("empty")
            grid_state.append(row)
        return grid_state

# Función para que el usuario elija el número de ticks
def get_user_ticks():
    while True:
        ticks = input("Ingresa el número de ticks para la simulación (por ejemplo, 50): ")
        try:
            ticks = int(ticks)
            if ticks > 0:
                return ticks
            else:
                print("Por favor, ingresa un número mayor que 0.")
        except ValueError:
            print("Entrada inválida. Intenta nuevamente.")

# Función para que el usuario elija el número de entidades iniciales
def get_user_entities():
    while True:
        try:
            num_plants = int(input("Ingresa el número inicial de plantas (1-3): "))
            num_animals = int(input("Ingresa el número inicial de animales (1-3): "))
            num_fungi = int(input("Ingresa el número inicial de hongos (1-3): "))

            if 1 <= num_plants <= 3 and 1 <= num_animals <= 3 and 1 <= num_fungi <= 3:
                return num_plants, num_animals, num_fungi
            else:
                print("Por favor, ingresa un número entre 1 y 3 para cada tipo de entidad.")
        except ValueError:
            print("Entrada inválida. Intenta nuevamente.")

# 🌍 **Inicialización del Mundo**
world = World(24, 24)

# 🎬 **Elección de Ticks y Entidades por Consola**
ticks = get_user_ticks()  # El usuario elige el número de ticks
num_plants, num_animals, num_fungi = get_user_entities()  # El usuario elige el número de entidades

# Generar entidades iniciales
for _ in range(num_plants):
    while True:
        x, y = random.randint(0, 23), random.randint(0, 23)
        if world.grid[y][x] == "soil" or world.grid[y][x] == "soil+":
            world.add_entity(PlantLow(x, y))
            break

for _ in range(num_animals):
    while True:
        x, y = random.randint(0, 23), random.randint(0, 23)
        if world.grid[y][x] is None:
            world.add_entity(AnimalSmall(x, y))
            break

for _ in range(num_fungi):
    while True:
        x, y = random.randint(0, 23), random.randint(0, 23)
        if world.grid[y][x] is None:
            world.add_entity(Fungi(x, y))
            break

# Crear un archivo Excel para guardar el resumen
workbook = Workbook()
sheet = workbook.active
sheet.title = "Tick 0"  # Hoja inicial

# Guardar el estado inicial de la cuadrícula
grid_state = world.get_grid_state()
for y in range(24):
    for x in range(24):
        sheet.cell(row=y + 1, column=x + 1, value=grid_state[y][x])

current_tick = 0

# 🎬 **Bucle Principal de Pygame**
running = True
while running:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False

    # Actualizar y dibujar el mundo
    if current_tick < ticks:
        world.update()
        world.display()
        current_tick += 1

        # Guardar el estado de la cuadrícula en una nueva hoja
        sheet = workbook.create_sheet(title=f"Tick {current_tick}")
        grid_state = world.get_grid_state()
        for y in range(24):
            for x in range(24):
                sheet.cell(row=y + 1, column=x + 1, value=grid_state[y][x])
    else:
        running = False  # Detener la simulación después de los ticks especificados

    # Controlar la velocidad de la simulación
    clock.tick(1)  # 1 tick por segundo

# Guardar el archivo Excel
workbook.save("simulation_summary.xlsx")
print("Resumen de la simulación guardado en 'simulation_summary.xlsx'.")

# Finalizar la simulación y cerrar Pygame
pygame.quit()